"""
Main Flask application for Teller Home App.
"""
import os
import logging
import json
import csv
import io
from flask import Flask, jsonify, render_template, request
from flask_cors import CORS
from dotenv import load_dotenv
from src.teller_client import TellerClient
from src.models import init_database, get_session, Account, Balance, Transaction, ScheduledPayment, UserEnrollment
from src.sync_service import SyncService
from datetime import datetime, timedelta
from sqlalchemy import desc
from openpyxl import load_workbook

load_dotenv()

# Configure logging - suppress SSL handshake errors
logging.basicConfig(level=logging.INFO)
werkzeug_logger = logging.getLogger('werkzeug')

# Filter out SSL/TLS handshake errors (code 400 with binary data)
class NoSSLFilter(logging.Filter):
    def filter(self, record):
        # Filter out SSL handshake attempts (binary data in logs)
        if hasattr(record, 'getMessage'):
            msg = record.getMessage()
            if 'Bad request syntax' in msg or 'Bad request version' in msg:
                if '\\x16\\x03\\x01' in msg or '\x16\x03\x01' in repr(msg):
                    return False
        return True

werkzeug_logger.addFilter(NoSSLFilter())

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
CORS(app)

# Initialize database
init_database()


@app.route('/')
def index():
    """Home page - redirect to static index."""
    from flask import send_from_directory
    import os
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    return send_from_directory(static_dir, 'index.html')


@app.route('/api/info')
def api_info():
    """API documentation endpoint."""
    return jsonify({
        "message": "Teller Home App API",
        "version": "1.0.0",
        "endpoints": {
            "/": "Home page with navigation",
            "/static/dashboard.html": "Dashboard with account cards",
            "/static/calendar.html": "Calendar for payment scheduling",
            "/static/transactions.html": "Transaction history and management",
            "/static/teller-connect.html": "Bank account connection UI",
            "/api/health": "Health check",
            "/api/sync": "Sync data from Teller",
            "/api/accounts": "Get all accounts",
            "/api/accounts/<id>": "Get specific account",
            "/api/accounts/<id>/balance": "Get current balance for account",
            "/api/accounts/<id>/transactions": "Get transactions for account",
            "/api/transactions": "Get all transactions with filtering, sorting, pagination",
            "/api/transactions/export": "Export transactions to CSV or JSON",
            "/api/scheduled-payments": "Get/Create scheduled payments",
            "/api/scheduled-payments/<id>": "Delete scheduled payment",
            "/api/scheduled-payments/import": "Import subscriptions from JSON/CSV file",
            "/api/weekly-forecast": "Get weekly financial forecast",
            "/api/teller-connect/enroll": "Enroll user with Teller Connect",
            "/api/teller-connect/status": "Get enrollment status",
            "/api/teller-connect/disconnect/<id>": "Disconnect enrollment"
        }
    })


@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files."""
    from flask import send_from_directory
    import os
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    return send_from_directory(static_dir, filename)


@app.route('/api/health')
def health():
    """Health check endpoint."""
    try:
        client = TellerClient()
        teller_status = "connected" if client.test_connection() else "disconnected"
    except Exception as e:
        teller_status = f"error: {str(e)}"
    
    return jsonify({
        "status": "ok",
        "timestamp": datetime.utcnow().isoformat(),
        "teller_api": teller_status
    })


@app.route('/api/sync', methods=['POST'])
def sync_data():
    """Trigger a sync from Teller API."""
    session = get_session()

    try:
        # Get the first active enrollment token
        enrollment = session.query(UserEnrollment).filter_by(is_active=True).first()

        if not enrollment:
            return jsonify({
                "status": "error",
                "message": "No active enrollments found. Please connect a bank account first."
            }), 400

        # Use the enrollment's access token
        client = TellerClient(app_token=enrollment.access_token)
        sync_service = SyncService(client, session)
        result = sync_service.sync_all()

        return jsonify({
            "status": "success",
            "synced": result,
            "enrollment_id": enrollment.enrollment_id,
            "timestamp": datetime.utcnow().isoformat()
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        session.close()


@app.route('/api/accounts')
def get_accounts():
    """Get all accounts, optionally filtered by budget persona."""
    session = get_session()

    try:
        budget = request.args.get('budget')
        query = session.query(Account)

        if budget == 'unassigned':
            query = query.filter(Account.budget_id == None)
        elif budget in ('dad', 'mom', 'house'):
            query = query.filter(Account.budget_id == budget)

        accounts = query.all()

        accounts_data = []
        for account in accounts:
            # Get latest balance
            latest_balance = session.query(Balance).filter_by(
                account_id=account.id
            ).order_by(desc(Balance.timestamp)).first()

            # Use ledger balance or available balance
            current_balance = latest_balance.ledger if latest_balance else 0

            accounts_data.append({
                "id": account.id,
                "name": account.name,
                "display_name": account.display_name,
                "type": account.type,
                "subtype": account.subtype,
                "institution_name": account.institution_name,
                "institution": account.institution_name,
                "currency": account.currency,
                "status": account.status,
                "budget_id": account.budget_id,
                "current_balance": current_balance,
                "balance": {
                    "available": latest_balance.available if latest_balance else None,
                    "ledger": latest_balance.ledger if latest_balance else None,
                    "timestamp": latest_balance.timestamp.isoformat() if latest_balance else None
                }
            })

        return jsonify({"accounts": accounts_data})

    finally:
        session.close()


@app.route('/api/accounts/<account_id>/budget', methods=['PUT'])
def update_account_budget(account_id):
    """Assign an account to a budget persona."""
    session = get_session()

    try:
        account = session.query(Account).filter_by(id=account_id).first()

        if not account:
            return jsonify({"error": "Account not found"}), 404

        data = request.get_json()
        budget_id = data.get('budget_id')

        if budget_id is not None and budget_id not in ('dad', 'mom', 'house'):
            return jsonify({"error": "budget_id must be 'dad', 'mom', 'house', or null"}), 400

        account.budget_id = budget_id
        account.updated_at = datetime.utcnow()
        session.commit()

        return jsonify({
            "status": "success",
            "account_id": account_id,
            "budget_id": account.budget_id
        })

    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        session.close()


@app.route('/api/budgets/summary')
def budgets_summary():
    """Return aggregate stats for each budget persona."""
    session = get_session()

    try:
        result = {}

        for persona in ('dad', 'mom', 'house'):
            accounts = session.query(Account).filter(Account.budget_id == persona).all()
            assets = 0.0
            liabilities = 0.0
            for account in accounts:
                latest_balance = session.query(Balance).filter_by(
                    account_id=account.id
                ).order_by(desc(Balance.timestamp)).first()
                balance = latest_balance.ledger if latest_balance else 0.0
                if account.type in ('credit_card', 'credit'):
                    liabilities += abs(balance)
                else:
                    assets += balance
            result[persona] = {
                "net_worth": round(assets - liabilities, 2),
                "assets": round(assets, 2),
                "liabilities": round(liabilities, 2),
                "account_count": len(accounts)
            }

        unassigned_count = session.query(Account).filter(Account.budget_id == None).count()
        result['unassigned'] = {"account_count": unassigned_count}

        return jsonify(result)

    finally:
        session.close()


@app.route('/api/accounts/<account_id>/display-name', methods=['PUT'])
def update_account_display_name(account_id):
    """Update custom display name for an account."""
    session = get_session()
    
    try:
        account = session.query(Account).filter_by(id=account_id).first()
        
        if not account:
            return jsonify({"error": "Account not found"}), 404
        
        data = request.get_json()
        display_name = data.get('display_name', '').strip()
        
        # Allow empty string to clear display name
        account.display_name = display_name if display_name else None
        account.updated_at = datetime.utcnow()
        session.commit()
        
        return jsonify({
            "status": "success",
            "account_id": account_id,
            "display_name": account.display_name
        })
    
    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()


@app.route('/api/accounts/<account_id>')
def get_account(account_id):
    """Get a specific account."""
    session = get_session()
    
    try:
        account = session.query(Account).filter_by(id=account_id).first()
        
        if not account:
            return jsonify({"error": "Account not found"}), 404
        
        # Get latest balance
        latest_balance = session.query(Balance).filter_by(
            account_id=account.id
        ).order_by(desc(Balance.timestamp)).first()
        
        return jsonify({
            "id": account.id,
            "name": account.name,
            "type": account.type,
            "subtype": account.subtype,
            "institution": account.institution_name,
            "currency": account.currency,
            "status": account.status,
            "balance": {
                "available": latest_balance.available if latest_balance else None,
                "ledger": latest_balance.ledger if latest_balance else None,
                "timestamp": latest_balance.timestamp.isoformat() if latest_balance else None
            },
            "created_at": account.created_at.isoformat(),
            "updated_at": account.updated_at.isoformat()
        })
    
    finally:
        session.close()


@app.route('/api/accounts/<account_id>/transactions')
def get_transactions(account_id):
    """Get transactions for a specific account."""
    session = get_session()
    
    try:
        limit = request.args.get('limit', 100, type=int)
        
        transactions = session.query(Transaction).filter_by(
            account_id=account_id
        ).order_by(desc(Transaction.date)).limit(limit).all()
        
        return jsonify([{
            "id": txn.id,
            "account_id": txn.account_id,
            "amount": txn.amount,
            "date": txn.date.isoformat(),
            "description": txn.description,
            "category": txn.category,
            "type": txn.type,
            "status": txn.status
        } for txn in transactions])

    finally:
        session.close()


@app.route('/api/transactions')
def get_all_transactions():
    """
    Get all transactions with filtering, sorting, and pagination.
    
    Query Parameters:
        account_id: Filter by specific account
        start_date: Filter transactions after date (ISO format)
        end_date: Filter transactions before date (ISO format)
        min_amount: Minimum transaction amount
        max_amount: Maximum transaction amount
        search: Search in description field
        sort_by: Sort field (date, amount, description) - default: date
        sort_order: asc or desc - default: desc
        page: Page number - default: 1
        per_page: Items per page - default: 50, max: 200
    """
    session = get_session()
    
    try:
        # Get query parameters
        account_id = request.args.get('account_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        min_amount = request.args.get('min_amount', type=float)
        max_amount = request.args.get('max_amount', type=float)
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'date')
        sort_order = request.args.get('sort_order', 'desc')
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 50, type=int), 200)
        
        # Build query
        query = session.query(Transaction)
        
        # Apply filters
        if account_id:
            query = query.filter(Transaction.account_id == account_id)
        
        if start_date:
            try:
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Transaction.date >= start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(Transaction.date <= end_dt)
            except ValueError:
                pass
        
        if min_amount is not None:
            query = query.filter(Transaction.amount >= min_amount)
        
        if max_amount is not None:
            query = query.filter(Transaction.amount <= max_amount)
        
        if search:
            query = query.filter(Transaction.description.ilike(f'%{search}%'))
        
        # Apply sorting
        sort_column = Transaction.date
        if sort_by == 'amount':
            sort_column = Transaction.amount
        elif sort_by == 'description':
            sort_column = Transaction.description
        
        if sort_order == 'asc':
            query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(sort_column.desc())
        
        # Get total count before pagination
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * per_page
        transactions = query.offset(offset).limit(per_page).all()
        
        # Get account names for display
        account_names = {}
        for txn in transactions:
            if txn.account_id not in account_names:
                account = session.query(Account).filter_by(id=txn.account_id).first()
                if account:
                    account_names[txn.account_id] = account.display_name or account.name
        
        # Format response
        transactions_data = [{
            "id": txn.id,
            "account_id": txn.account_id,
            "account_name": account_names.get(txn.account_id, 'Unknown'),
            "amount": txn.amount,
            "date": txn.date.isoformat(),
            "description": txn.description,
            "category": txn.category,
            "type": txn.type,
            "status": txn.status
        } for txn in transactions]
        
        # Calculate pagination info
        total_pages = (total_count + per_page - 1) // per_page
        
        return jsonify({
            "transactions": transactions_data,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total_count,
                "pages": total_pages
            },
            "filters_applied": {
                "account_id": account_id,
                "start_date": start_date,
                "end_date": end_date,
                "min_amount": min_amount,
                "max_amount": max_amount,
                "search": search,
                "sort_by": sort_by,
                "sort_order": sort_order
            }
        })
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()


@app.route('/api/transactions/export')
def export_transactions():
    """
    Export transactions to CSV or JSON format.
    Accepts same query parameters as /api/transactions.
    """
    session = get_session()
    
    try:
        export_format = request.args.get('format', 'csv').lower()
        
        # Get query parameters (same as get_all_transactions)
        account_id = request.args.get('account_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        min_amount = request.args.get('min_amount', type=float)
        max_amount = request.args.get('max_amount', type=float)
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'date')
        sort_order = request.args.get('sort_order', 'desc')
        
        # Build query (same logic as get_all_transactions)
        query = session.query(Transaction)
        
        if account_id:
            query = query.filter(Transaction.account_id == account_id)
        
        if start_date:
            try:
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Transaction.date >= start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(Transaction.date <= end_dt)
            except ValueError:
                pass
        
        if min_amount is not None:
            query = query.filter(Transaction.amount >= min_amount)
        
        if max_amount is not None:
            query = query.filter(Transaction.amount <= max_amount)
        
        if search:
            query = query.filter(Transaction.description.ilike(f'%{search}%'))
        
        # Apply sorting
        sort_column = Transaction.date
        if sort_by == 'amount':
            sort_column = Transaction.amount
        elif sort_by == 'description':
            sort_column = Transaction.description
        
        if sort_order == 'asc':
            query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(sort_column.desc())
        
        # Get all transactions (no pagination for export)
        transactions = query.all()
        
        # Get account names
        account_names = {}
        for txn in transactions:
            if txn.account_id not in account_names:
                account = session.query(Account).filter_by(id=txn.account_id).first()
                if account:
                    account_names[txn.account_id] = account.display_name or account.name
        
        if export_format == 'json':
            # Export as JSON
            transactions_data = [{
                "id": txn.id,
                "account_id": txn.account_id,
                "account_name": account_names.get(txn.account_id, 'Unknown'),
                "amount": txn.amount,
                "date": txn.date.isoformat(),
                "description": txn.description,
                "category": txn.category,
                "type": txn.type,
                "status": txn.status
            } for txn in transactions]
            
            response = jsonify(transactions_data)
            response.headers['Content-Disposition'] = f'attachment; filename=transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
            return response
        
        else:
            # Export as CSV
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['Date', 'Account', 'Description', 'Amount', 'Category', 'Type', 'Status'])
            
            # Write data
            for txn in transactions:
                writer.writerow([
                    txn.date.strftime('%Y-%m-%d %H:%M:%S'),
                    account_names.get(txn.account_id, 'Unknown'),
                    txn.description,
                    txn.amount,
                    txn.category or '',
                    txn.type or '',
                    txn.status
                ])
            
            output.seek(0)
            response = app.response_class(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
            )
            return response
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()
        session.close()


@app.route('/api/scheduled-payments', methods=['GET', 'POST'])
def scheduled_payments():
    """Get all scheduled payments or create a new one."""
    session = get_session()
    
    try:
        if request.method == 'GET':
            payments = session.query(ScheduledPayment).filter_by(is_active=True).all()

            return jsonify({
                "payments": [{
                    "id": payment.id,
                    "name": payment.name,
                    "amount": payment.amount,
                    "account_id": payment.account_id,
                    "day_of_month": payment.day_of_month,
                    "frequency": payment.frequency,
                    "email": payment.email,
                    "category": payment.category,
                    "notes": payment.notes,
                    "is_recurring": payment.is_recurring
                } for payment in payments],
                "count": len(payments)
            })

        else:  # POST
            data = request.json

            payment = ScheduledPayment(
                name=data['name'],
                amount=float(data['amount']),
                account_id=data.get('account_id'),
                day_of_month=int(data['day_of_month']),
                frequency=data.get('frequency', 'monthly'),
                email=data.get('email'),
                category=data.get('category'),
                notes=data.get('notes'),
                is_recurring=data.get('is_recurring', True)
            )

            session.add(payment)
            session.commit()

            return jsonify({
                "status": "success",
                "id": payment.id
            }), 201
    
    finally:
        session.close()


@app.route('/api/scheduled-payments/<int:payment_id>', methods=['DELETE'])
def delete_scheduled_payment(payment_id):
    """Delete a scheduled payment."""
    session = get_session()
    
    try:
        payment = session.query(ScheduledPayment).filter_by(id=payment_id).first()
        
        if not payment:
            return jsonify({"error": "Payment not found"}), 404
        
        # Soft delete
        payment.is_active = False
        session.commit()
        
        return jsonify({
            "status": "success",
            "message": f"Payment {payment_id} deleted"
        })
    
    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()


@app.route('/api/scheduled-payments/import', methods=['POST'])
def import_scheduled_payments():
    """
    Import scheduled payments from JSON or CSV file.

    Expected JSON format:
    [
        {
            "name": "Netflix",
            "email": "user@example.com",
            "amount": 15.99,
            "day_of_month": 1,
            "frequency": "monthly"
        }
    ]

    Expected CSV format (with header):
    name,email,amount,day_of_month,frequency
    Netflix,user@example.com,15.99,1,monthly
    """
    session = get_session()

    try:
        # Check if file upload or JSON body
        if 'file' in request.files:
            file = request.files['file']

            if not file or file.filename == '':
                return jsonify({"error": "No file provided"}), 400

            filename = file.filename.lower()

            # Parse based on file extension
            if filename.endswith('.json'):
                try:
                    file_content = file.read().decode('utf-8')
                    payments_data = json.loads(file_content)
                except json.JSONDecodeError as e:
                    return jsonify({"error": f"Invalid JSON format: {str(e)}"}), 400

            elif filename.endswith('.csv'):
                try:
                    file_content = file.read().decode('utf-8')
                    csv_reader = csv.DictReader(io.StringIO(file_content))
                    payments_data = list(csv_reader)
                except Exception as e:
                    return jsonify({"error": f"Invalid CSV format: {str(e)}"}), 400

            elif filename.endswith('.xlsx'):
                try:
                    # Load Excel file from bytes
                    workbook = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
                    sheet = workbook.active

                    # Get header row
                    headers = [cell.value for cell in sheet[1]]

                    # Parse rows into dictionaries
                    payments_data = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(row):  # Skip empty rows
                            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                            payments_data.append(row_dict)

                    workbook.close()
                except Exception as e:
                    return jsonify({"error": f"Invalid XLSX format: {str(e)}"}), 400

            else:
                return jsonify({"error": "File must be .json, .csv, or .xlsx"}), 400

        elif request.json:
            # Direct JSON body
            payments_data = request.json
            if not isinstance(payments_data, list):
                payments_data = [payments_data]

        else:
            return jsonify({"error": "No file or JSON data provided"}), 400

        # Validate and import payments
        imported = []
        errors = []

        for idx, payment_data in enumerate(payments_data):
            try:
                # Validate required fields
                required_fields = ['name', 'amount', 'day_of_month']
                missing_fields = [f for f in required_fields if f not in payment_data or not payment_data[f]]

                if missing_fields:
                    errors.append({
                        "row": idx + 1,
                        "error": f"Missing required fields: {', '.join(missing_fields)}"
                    })
                    continue

                # Create payment record
                payment = ScheduledPayment(
                    name=payment_data['name'],
                    amount=float(payment_data['amount']),
                    day_of_month=int(payment_data['day_of_month']),
                    frequency=payment_data.get('frequency', 'monthly').lower(),
                    email=payment_data.get('email'),
                    category=payment_data.get('category', 'subscription'),
                    notes=payment_data.get('notes'),
                    is_recurring=True,  # Subscriptions are always recurring
                    is_active=True
                )

                session.add(payment)
                imported.append({
                    "name": payment.name,
                    "amount": payment.amount,
                    "frequency": payment.frequency
                })

            except (ValueError, KeyError) as e:
                errors.append({
                    "row": idx + 1,
                    "error": str(e),
                    "data": payment_data
                })

        # Commit all successful imports
        if imported:
            session.commit()

        return jsonify({
            "status": "success" if not errors else "partial",
            "imported": len(imported),
            "errors": len(errors),
            "details": {
                "imported_payments": imported,
                "errors": errors
            }
        }), 200 if not errors else 207

    except Exception as e:
        session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        session.close()


@app.route('/api/weekly-forecast')
def weekly_forecast():
    """Get weekly financial forecast based on scheduled payments and current balances."""
    session = get_session()
    
    try:
        # Get all accounts with latest balances
        accounts = session.query(Account).all()
        total_balance = 0
        
        for account in accounts:
            latest_balance = session.query(Balance).filter_by(
                account_id=account.id
            ).order_by(desc(Balance.timestamp)).first()
            
            if latest_balance:
                total_balance += latest_balance.available
        
        # Get scheduled payments
        payments = session.query(ScheduledPayment).filter_by(is_active=True).all()
        
        # Generate 7-day forecast
        today = datetime.now()
        forecast = []
        
        for i in range(7):
            date = today + timedelta(days=i)
            day_of_month = date.day
            
            # Find payments due on this day
            day_payments = [p for p in payments if p.day_of_month == day_of_month]
            total_payments = sum(p.amount for p in day_payments)
            
            # Calculate projected balance
            projected_balance = total_balance - total_payments
            
            forecast.append({
                "date": date.strftime("%Y-%m-%d"),
                "day_name": date.strftime("%A"),
                "starting_balance": round(total_balance, 2),
                "payments": [{
                    "name": p.name,
                    "amount": p.amount,
                    "category": p.category
                } for p in day_payments],
                "total_payments": round(total_payments, 2),
                "ending_balance": round(projected_balance, 2)
            })
            
            # Update total_balance for next day
            total_balance = projected_balance
        
        return jsonify({
            "forecast": forecast,
            "generated_at": datetime.utcnow().isoformat()
        })
    
    finally:
        session.close()


@app.route('/api/teller-connect/enroll', methods=['POST'])
def enroll_teller():
    """
    Save Teller Connect enrollment and access token.
    Called after user completes Teller Connect flow.
    """
    data = request.json
    
    if not data.get('access_token') or not data.get('enrollment_id'):
        return jsonify({"error": "Missing required fields"}), 400
    
    session = get_session()
    try:
        # Check if enrollment already exists
        existing = session.query(UserEnrollment).filter_by(
            enrollment_id=data['enrollment_id']
        ).first()
        
        if existing:
            # Update existing enrollment
            existing.access_token = data['access_token']
            existing.is_active = True
            existing.updated_at = datetime.utcnow()
            if data.get('institution_name'):
                existing.institution_name = data['institution_name']
            enrollment = existing
        else:
            # Create new enrollment
            enrollment = UserEnrollment(
                enrollment_id=data['enrollment_id'],
                user_id=data.get('user_id', 'default_user'),
                access_token=data['access_token'],
                institution_name=data.get('institution_name'),
                created_at=datetime.utcnow()
            )
            session.add(enrollment)
        
        session.commit()
        
        # Immediately sync data for this user using their access token
        try:
            client = TellerClient(app_token=enrollment.access_token)
            sync_service = SyncService(client, session)
            result = sync_service.sync_all()
            
            # Update last_synced timestamp
            enrollment.last_synced = datetime.utcnow()
            session.commit()
            
            return jsonify({
                "status": "success",
                "enrollment_id": enrollment.enrollment_id,
                "user_id": enrollment.user_id,
                "synced": result,
                "message": "Successfully enrolled and synced accounts"
            })
        except Exception as sync_error:
            # Enrollment saved, but sync failed
            return jsonify({
                "status": "partial_success",
                "enrollment_id": enrollment.enrollment_id,
                "user_id": enrollment.user_id,
                "message": "Enrollment saved but sync failed",
                "error": str(sync_error)
            }), 207  # Multi-Status
    
    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()


@app.route('/api/teller-connect/status', methods=['GET'])
def teller_connect_status():
    """
    Get status of all Teller Connect enrollments.
    """
    user_id = request.args.get('user_id', 'default_user')
    
    session = get_session()
    try:
        enrollments = session.query(UserEnrollment).filter_by(
            user_id=user_id,
            is_active=True
        ).all()
        
        result = [{
            "enrollment_id": e.enrollment_id,
            "institution_name": e.institution_name,
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "last_synced": e.last_synced.isoformat() if e.last_synced else None,
            "is_active": e.is_active
        } for e in enrollments]
        
        return jsonify({
            "user_id": user_id,
            "enrollments": result,
            "count": len(result)
        })
    
    finally:
        session.close()


@app.route('/api/teller-connect/disconnect/<enrollment_id>', methods=['POST'])
def disconnect_enrollment(enrollment_id):
    """
    Deactivate a Teller Connect enrollment.
    """
    session = get_session()
    try:
        enrollment = session.query(UserEnrollment).filter_by(
            enrollment_id=enrollment_id
        ).first()
        
        if not enrollment:
            return jsonify({"error": "Enrollment not found"}), 404
        
        enrollment.is_active = False
        enrollment.updated_at = datetime.utcnow()
        session.commit()
        
        return jsonify({
            "status": "success",
            "message": f"Enrollment {enrollment_id} disconnected"
        })
    
    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500
    
    finally:
        session.close()


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5001, threaded=True)
